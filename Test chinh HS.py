#!/usr/bin/python3
########################################################### IMPORT MODULES - START #################################################################
from tkinter import *
from tkinter import messagebox
import time
from time import sleep, gmtime, strftime
from picamera import PiCamera
import cv2
import numpy as np
from tkinter import filedialog
from PIL import ImageTk, Image
import serial
from functools import partial
import math
from fractions import Fraction
from threading import *
import os
from tkinter import ttk
import awesometkinter as atk
import tkinter.font as font
import openpyxl
import subprocess
import shutil
import RPi.GPIO as GPIO
from ftplib import FTP
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Protection
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image as Img
from datetime import *
############################################################ IMPORT MODULES - END ##################################################################

########################################################## GLOBAL VARIABLE - START #################################################################
wait = 0
result_list = list(range(48))
covid19dir_old = ""
start_point = (0,0)
end_point = (0,0)
path = ''

hs = list(range(48))
workbook = openpyxl.load_workbook('/home/pi/Spotcheck/coefficient.xlsx')
sheet = workbook.active
for i in range(0,48):
    if(i<6):
        pos = str(chr(65+i+1)) + "2"
    if(i>=6 and i<12):
        pos = str(chr(65+i-5)) + "3"
    if(i>=12 and i<18):
        pos = str(chr(65+i-11)) + "4"
    if(i>=18 and i<24):
        pos = str(chr(65+i-17)) + "5"
    if(i>=24 and i<30):
        pos = str(chr(65+i-23)) + "6"
    if(i>=30 and i<36):
        pos = str(chr(65+i-29)) + "7"
    if(i>=36 and i<42):
        pos = str(chr(65+i-35)) + "8"
    if(i>=42):
        pos = str(chr(65+i-41)) + "9"
    hs[i] = float(sheet[pos].value)

fr1 = open("/home/pi/Spotcheck/coordinates.txt","r")
x1 = int(fr1.readline())
y1 = int(fr1.readline())
x2 = int(fr1.readline())
y2 = int(fr1.readline())
########################################################### GLOBAL VARIABLE - END ##################################################################

########################################################## MAIN WINDOW INIT - START ################################################################
root = Tk()
root.title(" ")
root.geometry('1024x600')
root.configure(background = "white")
root.attributes('-fullscreen', True)
root.resizable(False,False)
def disable_event():
    pass
root.protocol("WM_DELETE_WINDOW", disable_event)
s = ttk.Style()
s.theme_use('clam')
########################################################### MAIN WINDOW INIT - END #################################################################

############################################################ CAMERA INIT - START ###################################################################
def camera_capture(output):
    global stop_click
    stop_click = 0
    camera = PiCamera(framerate=Fraction(1,6), sensor_mode=3)
    camera.rotation = 180
    camera.iso = 200
    sleep(2)
    camera.shutter_speed = 6000000
    camera.exposure_mode = 'off'
    camera.capture(output)
    camera.close()
############################################################# CAMERA INIT - END ####################################################################

############################################################ SERIAL INIT - START ###################################################################
ser = serial.Serial(
    port = '/dev/serial0',
    baudrate = 115200,
    parity = serial.PARITY_NONE,
    stopbits = serial.STOPBITS_ONE,
    bytesize = serial.EIGHTBITS,
    timeout = 1
)
############################################################# SERIAL INIT - END ####################################################################

######################################################### SORTING CONTOURS - START #################################################################
def sorting_y(contour):
    rect_y = cv2.boundingRect(contour)
    return rect_y[1]
def sorting_x(contour):
    rect_x = cv2.boundingRect(contour)
    return rect_x[0]
def sorting_xy(contour):
    rect_xy = cv2.boundingRect(contour)
    return math.sqrt(math.pow(rect_xy[0],2) + math.pow(rect_xy[1],2))
########################################################## SORTING CONTOURS - END ##################################################################

########################################################## IMAGE ANALYSIS - START ##################################################################
def process_image(image_name, start_point=(x1,y1), end_point=(x2,y2)):
    image = cv2.imread(image_name)
    blur_img = cv2.GaussianBlur(image.copy(), (35,35), 0)
    gray_img = cv2.cvtColor(blur_img, cv2.COLOR_BGR2GRAY)
    thresh, binary_img = cv2.threshold(gray_img.copy(), 30, maxval=255, type=cv2.THRESH_BINARY)
    contours, hierarchy = cv2.findContours(binary_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    print("Number of contours: " + str(len(contours)))
    
    
    def sorting_y(contour):
        rect_y = cv2.boundingRect(contour)
        return rect_y[1]
    def sorting_x(contour):
        rect_x = cv2.boundingRect(contour)
        return rect_x[0]
    def sorting_xy(contour):
        rect_xy = cv2.boundingRect(contour)
        return math.sqrt(math.pow(rect_xy[0],2) + math.pow(rect_xy[1],2))
    
    

    contours = sorted(contours, key=lambda data:sorting_xy(data))

    contour_img = np.zeros_like(gray_img)
    contour_img = cv2.rectangle(contour_img, start_point, end_point, (255,255,255), -1)
    rect_w = end_point[0] - start_point[0]
    rect_h = end_point[1] - start_point[1]
    cell_w = round(rect_w/6)
    cell_h = round(rect_h/8)
    for i in range(1,6):
        contour_img = cv2.line(contour_img, (start_point[0]+i*cell_w,start_point[1]), (start_point[0]+i*cell_w,end_point[1]),(0,0,0), 4)
    for i in range(1,8):
        contour_img = cv2.line(contour_img, (start_point[0],start_point[1]+i*cell_h), (end_point[0],start_point[1]+i*cell_h),(0,0,0), 4)

    thresh1 , binary1_img = cv2.threshold(contour_img, 250, maxval=255, type=cv2.THRESH_BINARY)
    contours1, hierarchy1 = cv2.findContours(binary1_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)

    contours1 = sorted(contours1, key=lambda data:sorting_y(data))
    contours1_h1 = contours1[0:6]
    contours1_h2 = contours1[6:12]
    contours1_h3 = contours1[12:18]
    contours1_h4 = contours1[18:24]
    contours1_h5 = contours1[24:30]
    contours1_h6 = contours1[30:36]
    contours1_h7 = contours1[36:42]
    contours1_h8 = contours1[42:48]
    contours1_h1.sort(key=lambda data:sorting_x(data))
    contours1_h2.sort(key=lambda data:sorting_x(data))
    contours1_h3.sort(key=lambda data:sorting_x(data))
    contours1_h4.sort(key=lambda data:sorting_x(data))
    contours1_h5.sort(key=lambda data:sorting_x(data))
    contours1_h6.sort(key=lambda data:sorting_x(data))
    contours1_h7.sort(key=lambda data:sorting_x(data))
    contours1_h8.sort(key=lambda data:sorting_x(data))

    sorted_contours1 = contours1_h1 + contours1_h2 + contours1_h3 + contours1_h4 + contours1_h5 + contours1_h6 + contours1_h7 + contours1_h8

    list_intensities = []
    sum_intensities = []
    result_list = list(range(48))
    area = list(range(48))


#Gray
#     tmp_list = list(range(48))
#     blur1_img = cv2.GaussianBlur(image.copy(), (25,25), 0)
#     grayprocess_img = cv2.cvtColor(blur1_img, cv2.COLOR_BGR2GRAY)
#     #cv2.imwrite("mau.jpg",grayprocess_img)
#     for i in range(len(sorted_contours1)):
#         cimg = np.zeros_like(gray_img)
#         cv2.drawContours(cimg, sorted_contours1, i, color = 255, thickness = -1)
#         pts = np.where(cimg == 255)
#         list_intensities.append(grayprocess_img[pts[0], pts[1]])
#         list_intensities[i].sort()
#         #print("list_intensities",str(i),":",list_intensities[i])
#         #print("value", str(i), " : ", list_intensities[i][len(list_intensities[i])-1])
#         sum_intensities.append(sum(list_intensities[i][len(list_intensities[i])-280:]))
#         #sum_intensities.append(sum(list_intensities[i][len(list_intensities[i])-240:]))
#         area[i]= cv2.contourArea(sorted_contours1[i])
#         #result_list[i] = sum_intensities[i]
#         tmp_list[i] = sum_intensities[i]/1000
#         #result_list[i] = round(tmp_list[i])
#         #result_list[i] = round(round(tmp_list[i],1)*1.5)
#         result_list[i] = round(tmp_list[i],1)

#BGR
    blur1_img = cv2.GaussianBlur(image.copy(), (25,25), 0)
    tmp_list = list(range(48))
    list_bgrvalue = []
    list_index = list(range(48))
    for i in range(len(sorted_contours1)):
        list_index[i] = []
        cimg = np.zeros_like(gray_img)
        cv2.drawContours(cimg, sorted_contours1, i, color = 255, thickness = -1)
        pts = np.where(cimg == 255)
        list_bgrvalue.append(blur1_img[pts[0], pts[1]])
        for j in range(len(list_bgrvalue[i])):
             list_index[i].append(round((list_bgrvalue[i][j][1]*3 + list_bgrvalue[i][j][2])))
        list_index[i].sort()
        list_intensities.append(sum(list_index[i][len(list_index[i])-250:]))
        area[i]= cv2.contourArea(sorted_contours1[i])
        tmp_list[i] = list_intensities[i]/1000
        result_list[i] = round(tmp_list[i],1)

#HSV
#     tmp_list = list(range(48))
#     #blur1_img = cv2.fastNlMeansDenoisingColored(image.copy(),None,9,9,7,19)
#     #cv2.imwrite("mau1.jpg",blur1_img)
#     blur1_img = cv2.GaussianBlur(image.copy(), (3,3), 0)
#     cv2.imwrite("mau.jpg",blur_img)
#     hsv_img = cv2.cvtColor(blur1_img, cv2.COLOR_BGR2HSV)
#     list_hsvvalue = []
#     list_index = list(range(48))
#     for i in range(len(sorted_contours1)):
#         list_index[i] = []
#         cimg = np.zeros_like(gray_img)
#         cv2.drawContours(cimg, sorted_contours1, i, color = 255, thickness = -1)
#         pts = np.where(cimg == 255)
#         list_hsvvalue.append(hsv_img[pts[0], pts[1]])
#         for j in range(len(list_hsvvalue[i])):
#             list_index[i].append(list_hsvvalue[i][j][2])
#         list_index[i].sort()
#         #print(len(list_index[i]))
#         list_intensities.append(sum(list_index[i][len(list_index[i])-250:]))
#         #area[i]= cv2.contourArea(sorted_contours1[i])
#         result_list[i] = list_intensities[i]
#         tmp_list[i] = list_intensities[i]/1000
#         result_list[i] = round(tmp_list[i])

#Nhân hệ số
    global hs
    for i in range(len(sorted_contours1)):
        result_list[i] = round(result_list[i]*hs[i],1)

#     for i in range(len(sorted_contours1)):
#         if(i==0):
#             result_list[i] = round(result_list[i]*1.03,1)
#         if(i==47):
#             result_list[i] = round(result_list[i]*0.92,1)
#         if(i==37 or i==38 or i==39 or i==40 or
#            i==43 or i==44 or i==45 or
#            i==10 or i==16 or i==22 or i==28 or i==34):
#             result_list[i] = round(result_list[i]*0.98,1)
#         if(i==36 or i==42 or i==46 or i==4 or
#            i==5 or i==11 or i==17 or i==23 or i==29 or i==35 or i==41):
#             result_list[i] = round(result_list[i]*0.96,1)
    for i in range(1,7):
        result_list[6*i+1]=round(result_list[6*i+1]*(1-(0.01*round(result_list[6*i]/70) + 0.01*round(result_list[6*i+2]/70) + 0.01*round(result_list[6*i-5]/70) + 0.01*round(result_list[6*i+7]/70) + 0.003*round(result_list[6*i-6]/70) + 0.003*round(result_list[6*i-4]/70) + 0.003*round(result_list[6*i+6]/70) + 0.003*round(result_list[6*i+8]/70))),1)
        result_list[6*i+2]=round(result_list[6*i+2]*(1-(0.01*round(result_list[6*i+1]/70) + 0.01*round(result_list[6*i+3]/70) + 0.01*round(result_list[6*i-4]/70) + 0.01*round(result_list[6*i+8]/70) + 0.003*round(result_list[6*i-5]/70) + 0.003*round(result_list[6*i-3]/70) + 0.003*round(result_list[6*i+7]/70) + 0.003*round(result_list[6*i+9]/70))),1)
        result_list[6*i+3]=round(result_list[6*i+3]*(1-(0.01*round(result_list[6*i+2]/70) + 0.01*round(result_list[6*i+4]/70) + 0.01*round(result_list[6*i-3]/70) + 0.01*round(result_list[6*i+9]/70) + 0.003*round(result_list[6*i-4]/70) + 0.003*round(result_list[6*i-2]/70) + 0.003*round(result_list[6*i+8]/70) + 0.003*round(result_list[6*i+10]/70))),1)
        result_list[6*i+4]=round(result_list[6*i+4]*(1-(0.01*round(result_list[6*i+3]/70) + 0.01*round(result_list[6*i+5]/70) + 0.01*round(result_list[6*i-2]/70) + 0.01*round(result_list[6*i+10]/70) + 0.003*round(result_list[6*i-3]/70) + 0.003*round(result_list[6*i-1]/70) + 0.003*round(result_list[6*i+9]/70) + 0.003*round(result_list[6*i+11]/70))),1)
        result_list[6*i]=round(result_list[6*i]*(1-(0.01*round(result_list[6*i+1]/70) + 0.01*round(result_list[6*i-6]/70) + 0.01*round(result_list[6*i+6]/70))),1)
        result_list[6*i+5]=round(result_list[6*i+5]*(1-(0.01*round(result_list[6*i+4]/70) + 0.01*round(result_list[6*i-1]/70) + 0.01*round(result_list[6*i+11]/70))),1)

    for i in range(1,5):
        result_list[i]=round(result_list[i]*(1-(0.01*round(result_list[i-1]/70) + 0.01*round(result_list[i+1]/70) + 0.01*round(result_list[i+6]/70))),1)
        result_list[i+42]=round(result_list[i+42]*(1-(0.01*round(result_list[i+41]/70) + 0.01*round(result_list[i+43]/70) + 0.01*round(result_list[i+36]/70))),1)

    result_list[0]=round(result_list[0]*(1-(0.01*round(result_list[1]/70) + 0.01*round(result_list[6]/70))),1)
    result_list[5]=round(result_list[5]*(1-(0.01*round(result_list[4]/70) + 0.01*round(result_list[11]/70))),1)
    result_list[42]=round(result_list[42]*(1-(0.01*round(result_list[43]/70) + 0.01*round(result_list[36]/70))),1)
    result_list[47]=round(result_list[47]*(1-(0.01*round(result_list[46]/70) + 0.01*round(result_list[41]/70))),1)

    for i in range(len(sorted_contours1)):
        if(result_list[i]>99):
            result_list[i]=99

    for i in range(len(sorted_contours1)):
        if ((i!=0) and ((i+1)%6==0)):
            print('%.1f'%(result_list[i]))
        else:
            print('%.1f'%(result_list[i]), end = ' | ')

    blurori_img = cv2.GaussianBlur(image.copy(), (25,25), 0)
    for i in range(len(sorted_contours1)):
        cv2.drawContours(blurori_img, sorted_contours1, i, (0,255,0), thickness = 2)
    return (result_list, blurori_img)
########################################################### IMAGE ANALYSIS - END ###################################################################

############################################################ MAIN SCREEN - START ###################################################################
def mainscreen():
    mainscreen_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    mainscreen_labelframe.place(x=0,y=0)

    def capture_click():
        capture_button.place_forget()
        exit_button.place_forget()

        send_data = 'P'
        ser.write(send_data.encode())

        folder_name = strftime("TEST %m-%d-%Y %H.%M.%S")
        path0= os.path.join("/home/pi/Desktop/Test/", folder_name + "/")
        os.mkdir(path0)

        process_label = Label(mainscreen_labelframe, text='Đang xử lý...', bg='white', font=("Courier",13,'bold'))
        process_label.place(x=330,y=350)
        root.update_idletasks()
        scanposition_progressbar = ttk.Progressbar(mainscreen_labelframe, orient = HORIZONTAL, style="green.Horizontal.TProgressbar", length = 200, mode = 'determinate')
        scanposition_progressbar.place(x=290,y=310)
        scanposition_progressbar['value'] = 5
        root.update_idletasks()

        if(ser.in_waiting>0):
            receive_data = ser.readline().decode('utf-8').rstrip()
            print("Data received:", receive_data)
            scanposition_progressbar['value'] = 20
            root.update_idletasks()
            if(receive_data=='C'):
                global wait
                wait = 1


        while(wait!=1):
            root.update_idletasks()
            if(ser.in_waiting>0):
                receive_data = ser.readline().decode('utf-8').rstrip()
                print("Data received:", receive_data)
                if(receive_data=='C'):
                    wait = 1
                    break;
        while(wait==1):
            try:
                camera_capture(path0 + "anh-chup.jpg")
            except Exception as e :
                error = messagebox.askquestion("Lỗi: "+ str(e), "Bạn có muốn thoát chương trình ?", icon = "error")
                if(error=='yes'):
                    root.destroy()

            global result_list

            result_list, result_img = process_image(path0 + "anh-chup.jpg")
            cv2.imwrite(path0 + 'anh-xu-ly.jpg', result_img)

            workbook = Workbook()
            sheet = workbook.active
            sheet["A2"] = "A"
            sheet["A3"] = "B"
            sheet["A4"] = "C"
            sheet["A5"] = "D"
            sheet["A6"] = "E"
            sheet["A7"] = "F"
            sheet["A8"] = "G"
            sheet["A9"] = "H"
            sheet["B1"] = "1"
            sheet["C1"] = "2"
            sheet["D1"] = "3"
            sheet["E1"] = "4"
            sheet["F1"] = "5"
            sheet["G1"] = "6"
            for i in range(0,48):
                if(i<6):
                    pos = str(chr(65+i+1)) + "2"
                if(i>=6 and i<12):
                    pos = str(chr(65+i-5)) + "3"
                if(i>=12 and i<18):
                    pos = str(chr(65+i-11)) + "4"
                if(i>=18 and i<24):
                    pos = str(chr(65+i-17)) + "5"
                if(i>=24 and i<30):
                    pos = str(chr(65+i-23)) + "6"
                if(i>=30 and i<36):
                    pos = str(chr(65+i-29)) + "7"
                if(i>=36 and i<42):
                    pos = str(chr(65+i-35)) + "8"
                if(i>=42):
                    pos = str(chr(65+i-41)) + "9"

                sheet[pos] = result_list[i]

            workbook.save(path0 + "gia-tri.xlsx")

            scanposition_progressbar['value'] = 50
            root.update_idletasks()
            sleep(1)
            scanposition_progressbar['value'] = 70
            root.update_idletasks()
            sleep(1)
            scanposition_progressbar['value'] = 99
            root.update_idletasks()

            process_label.place_forget()
            scanposition_progressbar.place_forget()
            wait = 0

            result_labelframe = LabelFrame(mainscreen_labelframe, bg='ghost white', width=600,height = 307)
            result_labelframe.place(x=5,y=30)

            a1_labelframe = LabelFrame(mainscreen_labelframe, bg='white', width=250, height=339)
            a1_labelframe.place(x=269,y=30)
            a1 = Image.open(path0 + 'anh-chup.jpg')
            a1_crop = a1.crop((x1-5, y1-5, x2+5, y2+5))
            crop_width, crop_height = a1_crop.size
            scale_percent = 100
            width = int(crop_width * scale_percent / 100)
            height = int(crop_height * scale_percent / 100)
            display_img = a1_crop.resize((width,height))
            a1_display = ImageTk.PhotoImage(display_img)
            a1_label = Label(a1_labelframe, image=a1_display)
            a1_label.image = a1_display
            a1_label.place(x=0,y=1)

            a2_labelframe = LabelFrame(mainscreen_labelframe, bg='white', width=250, height=339)
            a2_labelframe.place(x=540,y=30)
            a2 = Image.open(path0 + 'anh-xu-ly.jpg')
            a2_crop = a2.crop((x1-5, y1-5, x2+5, y2+5))
            crop_width, crop_height = a2_crop.size
            scale_percent = 100
            width = int(crop_width * scale_percent / 100)
            height = int(crop_height * scale_percent / 100)
            display_img = a2_crop.resize((width,height))
            a2_display = ImageTk.PhotoImage(display_img)
            a2_label = Label(a2_labelframe, image=a2_display)
            a2_label.image = a2_display
            a2_label.place(x=0,y=1)

            root.update_idletasks()

            label = list(range(48))
            def result_table(range_a,range_b, row_value):
                j=-1
                for i in range(range_a, range_b):
                    j+=1
                    label[i] = Label(result_labelframe, bg='lawn green', text=str(result_list[i]), width=4, height=2)
                    label[i].grid(row=row_value,column=j,padx=2,pady=2)

            result_table(0,6,0)
            result_table(6,12,1)
            result_table(12,18,2)
            result_table(18,24,3)
            result_table(24,30,4)
            result_table(30,36,5)
            result_table(36,42,6)
            result_table(42,48,7)
            root.update_idletasks()

            def back_click():
                back_button.place_forget()
                result_labelframe.place_forget()
                a1_labelframe.place_forget()
                a2_labelframe.place_forget()
                mainscreen()

            back_button = Button(mainscreen_labelframe, bg="dark orange", text="Back", height=3, width=15, borderwidth=0, command=back_click)
            back_button.place(x=325,y=410)

    def exit_click():
        msg = messagebox.askquestion('','Thoát hả bro?', icon = 'question')
        if(msg=='yes'):
            root.destroy()


    capture_button = Button(mainscreen_labelframe, bg="dodger blue", activebackground="white", text="CAPTURE", font=("Courier",14,'bold'), borderwidth=0, height=3, width=15,command=capture_click)
    capture_button.place(x=450,y=300)
    exit_button = Button(mainscreen_labelframe, bg="red", activebackground="white", text="EXIT", font=("Courier",14,'bold'), borderwidth=0, height=3, width=15, command=exit_click)
    exit_button.place(x=150,y=300)

############################################################# MAIN SCREEN - END ####################################################################

############################################################### LOOP - START #######################################################################
while True:
    mainscreen()
    root.mainloop()
################################################################ LOOP - END ########################################################################
